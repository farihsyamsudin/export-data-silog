import os
import sys
import argparse
import pandas as pd
import psycopg2
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from sqlalchemy import create_engine

# =========================================================
# 🎯 Parse Command Line Arguments
# =========================================================
parser = argparse.ArgumentParser(description='Export Inventaris Data')
parser.add_argument('--polda-only', action='store_true', help='Export hanya data POLDA')
parser.add_argument('--polres-only', action='store_true', help='Export hanya data POLRES')
parser.add_argument('--polsek-only', action='store_true', help='Export hanya data POLSEK')
parser.add_argument('--satker-mabes-only', action='store_true', help='Export hanya data Satker Mabes')
args = parser.parse_args()

# Tentukan mode export
export_all = not (args.polda_only or args.polres_only or args.polsek_only or args.satker_mabes_only)
export_polda = export_all or args.polda_only
export_polres = export_all or args.polres_only
export_polsek = export_all or args.polsek_only
export_satker_mabes = export_all or args.satker_mabes_only

print("🎯 Mode Export:")
if export_all:
    print("   ➜ ALL (POLDA, POLRES, POLSEK, Satker Mabes)")
else:
    if export_polda: print("   ➜ POLDA")
    if export_polres: print("   ➜ POLRES")
    if export_polsek: print("   ➜ POLSEK")
    if export_satker_mabes: print("   ➜ Satker Mabes")
print()

# =========================================================
# 1️⃣ Load konfigurasi dari .env
# =========================================================
load_dotenv()

DB_HOST = os.getenv("DB_HOST")
DB_PORT = os.getenv("DB_PORT")
DB_NAME = os.getenv("DB_DATABASE")
DB_USER = os.getenv("DB_USERNAME")
DB_PASS = os.getenv("DB_PASSWORD")

engine = create_engine(f"postgresql://{DB_USER}:{DB_PASS}@{DB_HOST}:{DB_PORT}/{DB_NAME}")

# Direktori output utama
output_dir = "exports"
os.makedirs(output_dir, exist_ok=True)

# =========================================================
# 2️⃣ Fungsi bantu
# =========================================================
def style_header(ws):
    for row_num in [1, 2]:
        for cell in ws[row_num]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = "C3"

def auto_resize_columns(ws):
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if cell.value:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

def zero_to_empty(value):
    return "" if value == 0 else value

def sanitize_name(name):
    return name[:31].replace('/', '-').replace('\\', '-').replace('*', '').replace('?', '').replace(':', '').replace('[', '').replace(']', '')

# =========================================================
# 🏛️ FUNGSI UNTUK SATKER MABES
# =========================================================
def get_all_children_recursive(satker_id, all_satkers_df):
    """Mendapatkan semua child dari satker secara rekursif (depth-first)"""
    children = []
    direct_children = all_satkers_df[all_satkers_df['parent_id'] == satker_id].sort_values('name')
    
    for _, child in direct_children.iterrows():
        children.append(child)
        # Rekursif untuk mendapatkan anak-anak dari child ini
        grandchildren = get_all_children_recursive(child['id'], all_satkers_df)
        children.extend(grandchildren)
    
    return children

def get_parent_chain(satker_id, all_satkers_df):
    """Mendapatkan chain parent dari satker ke atas (untuk nama file)"""
    chain = []
    current_id = satker_id
    
    while current_id is not None:
        satker_match = all_satkers_df[all_satkers_df['id'] == current_id]
        
        # Jika tidak menemukan satker dengan ID tersebut, hentikan loop
        if satker_match.empty:
            print(f"    ⚠️ Warning: Satker dengan ID {current_id} tidak ditemukan")
            break
            
        satker = satker_match.iloc[0]
        chain.append(satker['name'])
        
        # Handle parent_id yang bisa None atau NaN
        parent_id = satker['parent_id']
        current_id = None if pd.isna(parent_id) else int(parent_id)
    
    # Balik urutan agar dari level tertinggi ke terendah
    return list(reversed(chain))

def export_satker_mabes():
    """Export data Satker Mabes dengan hierarki"""
    print("🏛️ Processing Satker Mabes...")
    
    # Buat folder satker_mabes
    satker_output_dir = os.path.join(output_dir, 'satker_mabes')
    os.makedirs(satker_output_dir, exist_ok=True)
    
    # Ambil semua satker mabes
    satkers_query = "SELECT id, name, level, parent_id FROM satker_mabes ORDER BY level, name;"
    df_all_satkers = pd.read_sql(satkers_query, engine)
    
    if df_all_satkers.empty:
        print("⚠️ Tidak ada data Satker Mabes")
        return
    
    # Process setiap satker
    for _, satker in df_all_satkers.iterrows():
        satker_id = satker['id']
        satker_name = satker['name']
        satker_level = satker['level']
        
        # Dapatkan parent chain untuk nama file
        parent_chain = get_parent_chain(satker_id, df_all_satkers)
        # Buat nama file sesuai level: Level1_Level2_Level3
        file_display_name = '_'.join(parent_chain)
        
        print(f"  -> Processing: {satker_name} (Level {satker_level}) -> File: {file_display_name}")
        
        # Dapatkan satker ini sendiri + semua children secara rekursif
        all_related_satkers = [satker]
        children = get_all_children_recursive(satker_id, df_all_satkers)
        all_related_satkers.extend(children)
        
        # Buat list nama untuk header
        satker_names = [s['name'] for s in all_related_satkers]
        
        # Query inventaris untuk satker ini dan semua children-nya
        satker_ids = [s['id'] for s in all_related_satkers]
        satker_ids_str = ','.join(map(str, satker_ids))
        
        inventory_query = f"""
            SELECT
                et.id AS penggolongan_id, et.name AS penggolongan,
                e.name AS jenis_materiil, e."order",
                inv.satker_name,
                COALESCE(inv.baik, 0) AS baik,
                COALESCE(inv.rusak_ringan, 0) AS rusak_ringan,
                COALESCE(inv.rusak_berat, 0) AS rusak_berat
            FROM equipments e
            JOIN equipment_types et ON et.id = e.id_equipment_type
            LEFT JOIN (
                SELECT
                    ei.equipment_id, sm.name AS satker_name,
                    SUM(ei.baik) AS baik, SUM(ei.rusak_ringan) AS rusak_ringan, SUM(ei.rusak_berat) AS rusak_berat
                FROM equipment_inventories ei
                JOIN satker_mabes sm ON sm.id = ei.owner_id
                WHERE ei.owner_type = 'App\\Models\\SatkerMabes' AND sm.id IN ({satker_ids_str})
                GROUP BY ei.equipment_id, sm.name
            ) AS inv ON e.id = inv.equipment_id
            WHERE e.deleted_at is null
            ORDER BY et.id, e."order";
        """
        
        df_inventory = pd.read_sql(inventory_query, engine)
        
        # Buat workbook
        wb = Workbook()
        ws = wb.active
        ws.title = sanitize_name(satker_name)
        
        # Header baris 1
        header1 = ["No.", "Jenis Materil"]
        for name in satker_names:
            header1 += [name, "", "", ""]
        ws.append(header1)
        
        # Merge cells untuk header satker
        for i, _ in enumerate(satker_names):
            start_col = 3 + (i * 4)
            ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=start_col + 3)
        
        # Header baris 2
        header2 = ["", ""]
        header2 += ["Baik", "Rusak Ringan", "Rusak Berat", "Jumlah"] * len(satker_names)
        ws.append(header2)
        
        current_row = 3
        for penggolongan, group_df in df_inventory.groupby("penggolongan", sort=False):
            ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=len(header1))
            ws.cell(row=current_row, column=2, value=penggolongan).font = Font(bold=True)
            current_row += 1
            
            for jenis_no, (jenis, jenis_df) in enumerate(group_df.groupby("jenis_materiil", sort=False), start=1):
                row_data = [jenis_no, jenis]
                for name in satker_names:
                    row = jenis_df[jenis_df["satker_name"] == name]
                    baik = int(row["baik"].iloc[0]) if not row.empty else 0
                    rr = int(row["rusak_ringan"].iloc[0]) if not row.empty else 0
                    rb = int(row["rusak_berat"].iloc[0]) if not row.empty else 0
                    
                    jumlah = baik + rr + rb
                    row_data += [zero_to_empty(baik), zero_to_empty(rr), zero_to_empty(rb), zero_to_empty(jumlah)]
                
                ws.append(row_data)
                ws.cell(row=current_row, column=1).alignment = Alignment(horizontal="center", vertical="center")
                current_row += 1
        
        style_header(ws)
        auto_resize_columns(ws)
        
        # Simpan file dengan nama sesuai hierarki
        filename = os.path.join(satker_output_dir, f"{sanitize_name(file_display_name)}.xlsx")
        wb.save(filename)
        print(f"    ✅ Saved: {filename}")
    
    print("✅ Satker Mabes export selesai!\n")

# =========================================================
# 3️⃣ EXPORT POLDA, POLRES, POLSEK (ENHANCED - SINGLE FILE)
# =========================================================
if export_polda or export_polres or export_polsek:
    poldas = pd.read_sql("SELECT id, name FROM polda ORDER BY id", engine)
    
    for _, polda in poldas.iterrows():
        polda_id = polda["id"]
        polda_name = polda["name"]
        
        print(f"🚀 Processing POLDA: {polda_name}")
        
        polda_output_dir = os.path.join(output_dir, 'POLDA ' + polda_name)
        os.makedirs(polda_output_dir, exist_ok=True)
        
        # Ambil daftar Subsatker dan Polres
        subsatkers_list_query = f"SELECT name FROM subsatker_poldas WHERE polda_id = {polda_id} ORDER BY name;"
        df_subsatkers_list = pd.read_sql(subsatkers_list_query, engine)
        subsatkers = df_subsatkers_list["name"].tolist()
        
        polres_list_query = f"SELECT id AS polres_id, name AS polres_name FROM polres WHERE polda_id = {polda_id} ORDER BY name;"
        df_polres_list = pd.read_sql(polres_list_query, engine)
        
        # Buat workbook untuk POLDA (single file)
        wb_polda = Workbook()
        ws_polda = wb_polda.active
        ws_polda.title = sanitize_name('POLDA ' + polda_name)
        
        # ===== SHEET POLDA =====
        if export_polda:
            subsatker_query = f"""
                SELECT
                    et.id AS penggolongan_id, et.name AS penggolongan,
                    e.name AS jenis_materiil, e."order",
                    inv.subsatker_name,
                    COALESCE(inv.baik, 0) AS baik,
                    COALESCE(inv.rusak_ringan, 0) AS rusak_ringan,
                    COALESCE(inv.rusak_berat, 0) AS rusak_berat
                FROM equipments e
                JOIN equipment_types et ON et.id = e.id_equipment_type
                LEFT JOIN (
                    SELECT
                        ei.equipment_id, sp.name AS subsatker_name,
                        SUM(ei.baik) AS baik, SUM(ei.rusak_ringan) AS rusak_ringan, SUM(ei.rusak_berat) AS rusak_berat
                    FROM equipment_inventories ei
                    JOIN subsatker_poldas sp ON sp.id = ei.owner_id
                    WHERE ei.owner_type = 'App\\Models\\SubsatkerPolda' AND sp.polda_id = {polda_id}
                    GROUP BY ei.equipment_id, sp.name
                ) AS inv ON e.id = inv.equipment_id
                WHERE e.deleted_at is null
                ORDER BY et.id, e."order";
            """
            df_subsatker = pd.read_sql(subsatker_query, engine)
            
            if not df_subsatker.empty:
                header1 = ["No.", "Jenis Materil"]
                for s in subsatkers:
                    header1 += [s, "", "", ""]
                ws_polda.append(header1)
                
                for i, _ in enumerate(subsatkers):
                    start_col = 3 + (i * 4)
                    ws_polda.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=start_col + 3)
                
                header2 = ["", ""]
                header2 += ["Baik", "Rusak Ringan", "Rusak Berat", "Jumlah"] * len(subsatkers)
                ws_polda.append(header2)
                
                current_row = 3
                for penggolongan, group_df in df_subsatker.groupby("penggolongan", sort=False):
                    ws_polda.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=len(header1))
                    ws_polda.cell(row=current_row, column=2, value=penggolongan).font = Font(bold=True)
                    current_row += 1
                    
                    for jenis_no, (jenis, jenis_df) in enumerate(group_df.groupby("jenis_materiil", sort=False), start=1):
                        row_data = [jenis_no, jenis]
                        for s in subsatkers:
                            row = jenis_df[jenis_df["subsatker_name"] == s]
                            baik = int(row["baik"].iloc[0]) if not row.empty else 0
                            rr = int(row["rusak_ringan"].iloc[0]) if not row.empty else 0
                            rb = int(row["rusak_berat"].iloc[0]) if not row.empty else 0
                            
                            jumlah_subsatker = baik + rr + rb
                            row_data += [zero_to_empty(baik), zero_to_empty(rr), zero_to_empty(rb), zero_to_empty(jumlah_subsatker)]
                        
                        ws_polda.append(row_data)
                        ws_polda.cell(row=current_row, column=1).alignment = Alignment(horizontal="center", vertical="center")
                        current_row += 1
                
                style_header(ws_polda)
                auto_resize_columns(ws_polda)
        
        # ===== SHEETS POLRES (dengan Polsek sebagai header horizontal) =====
        if export_polres:
            for _, polres_row in df_polres_list.iterrows():
                polres_id = polres_row["polres_id"]
                polres_name = polres_row["polres_name"]
                
                print(f"  -> Processing POLRES: {polres_name}")
                
                # Ambil daftar Polsek di bawah Polres ini
                polsek_list_query = f"SELECT id, name FROM polsek WHERE polres_id = {polres_id} ORDER BY name;"
                df_polsek_list = pd.read_sql(polsek_list_query, engine)
                
                # Buat list unit: POLRES + Polsek-polseknya
                units = [polres_name] + df_polsek_list["name"].tolist()
                
                # Query untuk mendapatkan SEMUA equipment dengan inventaris per unit
                polres_polsek_query = f"""
                    SELECT
                        et.id AS penggolongan_id, et.name AS penggolongan,
                        e.name AS jenis_materiil, e."order",
                        inv.unit_name,
                        COALESCE(inv.baik, 0) AS baik,
                        COALESCE(inv.rusak_ringan, 0) AS rusak_ringan,
                        COALESCE(inv.rusak_berat, 0) AS rusak_berat
                    FROM equipments e
                    JOIN equipment_types et ON et.id = e.id_equipment_type
                    LEFT JOIN (
                        SELECT
                            ei.equipment_id,
                            CASE 
                                WHEN ei.owner_type = 'App\\Models\\Polres' THEN p.name
                                WHEN ei.owner_type = 'App\\Models\\Polsek' THEN ps.name
                            END AS unit_name,
                            SUM(ei.baik) AS baik,
                            SUM(ei.rusak_ringan) AS rusak_ringan,
                            SUM(ei.rusak_berat) AS rusak_berat
                        FROM equipment_inventories ei
                        LEFT JOIN polres p ON ei.owner_type = 'App\\Models\\Polres' AND p.id = ei.owner_id AND p.id = {polres_id}
                        LEFT JOIN polsek ps ON ei.owner_type = 'App\\Models\\Polsek' AND ps.id = ei.owner_id AND ps.polres_id = {polres_id}
                        WHERE (
                            (ei.owner_type = 'App\\Models\\Polres' AND ei.owner_id = {polres_id})
                            OR (ei.owner_type = 'App\\Models\\Polsek' AND ps.polres_id = {polres_id})
                        )
                        GROUP BY ei.equipment_id, ei.owner_type, p.name, ps.name
                    ) AS inv ON e.id = inv.equipment_id
                    WHERE e.deleted_at IS NULL
                    ORDER BY et.id, e."order";
                """
                
                df_polres_polsek = pd.read_sql(polres_polsek_query, engine)
                
                if df_polres_polsek.empty:
                    continue
                
                # Buat sheet baru untuk Polres ini
                ws_polres = wb_polda.create_sheet(sanitize_name(polres_name))
                
                # Header baris 1 - nama unit
                header1 = ["No.", "Jenis Materil"]
                for unit in units:
                    header1 += [unit, "", "", ""]
                ws_polres.append(header1)
                
                # Merge cells untuk header unit
                for i, _ in enumerate(units):
                    start_col = 3 + (i * 4)
                    ws_polres.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=start_col + 3)
                
                # Header baris 2 - Baik, RR, RB, Jumlah
                header2 = ["", ""]
                header2 += ["Baik", "Rusak Ringan", "Rusak Berat", "Jumlah"] * len(units)
                ws_polres.append(header2)
                
                # Isi data
                current_row = 3
                for penggolongan, group_df in df_polres_polsek.groupby("penggolongan", sort=False):
                    ws_polres.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=len(header1))
                    ws_polres.cell(row=current_row, column=2, value=penggolongan).font = Font(bold=True)
                    current_row += 1
                    
                    for jenis_no, (jenis, jenis_df) in enumerate(group_df.groupby("jenis_materiil", sort=False), start=1):
                        row_data = [jenis_no, jenis]
                        
                        # Loop untuk setiap unit (POLRES + Polsek)
                        for unit in units:
                            row = jenis_df[jenis_df["unit_name"] == unit]
                            baik = int(row["baik"].iloc[0]) if not row.empty else 0
                            rr = int(row["rusak_ringan"].iloc[0]) if not row.empty else 0
                            rb = int(row["rusak_berat"].iloc[0]) if not row.empty else 0
                            
                            jumlah = baik + rr + rb
                            row_data += [zero_to_empty(baik), zero_to_empty(rr), zero_to_empty(rb), zero_to_empty(jumlah)]
                        
                        ws_polres.append(row_data)
                        ws_polres.cell(row=current_row, column=1).alignment = Alignment(horizontal="center", vertical="center")
                        current_row += 1
                
                style_header(ws_polres)
                auto_resize_columns(ws_polres)
        
        # Simpan file POLDA (single file dengan semua sheets)
        polda_filename = os.path.join(polda_output_dir, f"Inventaris_POLDA_{polda_name}.xlsx")
        wb_polda.save(polda_filename)
        print(f"✅ Saved {polda_filename}\n")

# =========================================================
# 4️⃣ EXPORT SATKER MABES
# =========================================================
if export_satker_mabes:
    export_satker_mabes()

print("\n🎉 Semua file selesai dibuat di folder 'exports'!")