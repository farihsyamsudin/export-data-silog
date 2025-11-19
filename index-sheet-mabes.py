import os
import sys
import argparse
import pandas as pd
import psycopg2
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from sqlalchemy import create_engine

# =========================================================
# 🎯 Parse Command Line Arguments
# =========================================================
parser = argparse.ArgumentParser(description='Export Inventaris Data')
parser.add_argument('--polda-only', action='store_true', help='Export hanya data POLDA')
parser.add_argument('--polres-only', action='store_true', help='Export hanya data POLRES')
parser.add_argument('--polsek-only', action='store_true', help='Export hanya data POLSEK')
parser.add_argument('--satker-mabes-only', action='store_true', help='Export hanya data Satker Mabes')
args = parser.parse_args()

# Tentukan mode export
export_all = not (args.polda_only or args.polres_only or args.polsek_only or args.satker_mabes_only)
export_polda = export_all or args.polda_only
export_polres = export_all or args.polres_only
export_polsek = export_all or args.polsek_only
export_satker_mabes = export_all or args.satker_mabes_only

print("🎯 Mode Export:")
if export_all:
    print("   ➜ ALL (POLDA, POLRES, POLSEK, Satker Mabes)")
else:
    if export_polda: print("   ➜ POLDA")
    if export_polres: print("   ➜ POLRES")
    if export_polsek: print("   ➜ POLSEK")
    if export_satker_mabes: print("   ➜ Satker Mabes")
print()

# =========================================================
# 1️⃣ Load konfigurasi dari .env
# =========================================================
load_dotenv()

DB_HOST = os.getenv("DB_HOST")
DB_PORT = os.getenv("DB_PORT")
DB_NAME = os.getenv("DB_DATABASE")
DB_USER = os.getenv("DB_USERNAME")
DB_PASS = os.getenv("DB_PASSWORD")

engine = create_engine(f"postgresql://{DB_USER}:{DB_PASS}@{DB_HOST}:{DB_PORT}/{DB_NAME}")

# Direktori output utama
output_dir = "exports"
os.makedirs(output_dir, exist_ok=True)

# =========================================================
# 2️⃣ Fungsi bantu
# =========================================================
def style_header_simple(ws):
    """Fungsi styling baru untuk header tunggal"""
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"

def auto_resize_columns(ws):
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if cell.value:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

def zero_to_empty(value):
    return "" if value == 0 else value

def sanitize_name(name):
    return (name
        .replace('/', '-')
        .replace('\\', '-')
        .replace('*', '')
        .replace('?', '')
        .replace(':', '')
        .replace('[', '')
        .replace(']', '')
        .replace('.', '')
        .strip()
    )

# =========================================================
# 🏛️ FUNGSI UNTUK SATKER MABES (TELAH DIPERBARUI)
# =========================================================
def get_all_children_recursive(satker_id, all_satkers_df):
    """Mendapatkan semua child dari satker secara rekursif (depth-first)"""
    children = []
    direct_children = all_satkers_df[all_satkers_df['parent_id'] == satker_id].sort_values('name')
    
    for _, child in direct_children.iterrows():
        children.append(child)
        grandchildren = get_all_children_recursive(child['id'], all_satkers_df)
        children.extend(grandchildren)
    
    return children

def get_parent_chain(satker_id, all_satkers_df):
    """Mendapatkan chain parent dari satker ke atas (untuk nama file)"""
    chain = []
    current_id = satker_id
    
    while current_id is not None:
        satker_match = all_satkers_df[all_satkers_df['id'] == current_id]
        if satker_match.empty:
            print(f"    ⚠️ Warning: Satker dengan ID {current_id} tidak ditemukan")
            break
        satker = satker_match.iloc[0]
        chain.append(satker['name'])
        parent_id = satker['parent_id']
        current_id = None if pd.isna(parent_id) else int(parent_id)
    return list(reversed(chain))

def export_satker_mabes():
    """Export data Satker Mabes dengan hierarki menjadi sheet, tanpa skip data kosong."""
    print("🏛️ Processing Satker Mabes...")
    
    satker_output_dir = os.path.join(output_dir, 'satker_mabes')
    os.makedirs(satker_output_dir, exist_ok=True)
    
    satkers_query = "SELECT id, name, level, parent_id FROM satker_mabes ORDER BY level, name;"
    df_all_satkers = pd.read_sql(satkers_query, engine)
    
    if df_all_satkers.empty:
        print("⚠️ Tidak ada data Satker Mabes")
        return
    
    for _, satker in df_all_satkers.iterrows():
        satker_id = satker['id']
        
        parent_chain = get_parent_chain(satker_id, df_all_satkers)
        file_display_name = '_'.join(parent_chain)
        
        print(f"  -> Processing File: {file_display_name}.xlsx")
        
        all_related_satkers = [satker]
        children = get_all_children_recursive(satker_id, df_all_satkers)
        all_related_satkers.extend(children)
        
        satker_ids = [s['id'] for s in all_related_satkers]
        if not satker_ids:
            continue
        satker_ids_str = ','.join(map(str, satker_ids))
        
        # 🔧 FIX: Tambahkan satker_id di SELECT untuk identifier unik
        inventory_query = f"""
            WITH relevant_satkers AS (
                SELECT id, name FROM satker_mabes WHERE id IN ({satker_ids_str})
            )
            SELECT
                et.id AS penggolongan_id,
                et.name AS penggolongan,
                e.name AS jenis_materiil,
                e."order",
                rs.id AS satker_id,
                rs.name AS satker_name,
                COALESCE(inv.baik, 0) AS baik,
                COALESCE(inv.rusak_ringan, 0) AS rusak_ringan,
                COALESCE(inv.rusak_berat, 0) AS rusak_berat
            FROM equipments e
            JOIN equipment_types et ON et.id = e.id_equipment_type
            CROSS JOIN relevant_satkers rs
            LEFT JOIN (
                SELECT
                    ei.equipment_id,
                    ei.owner_id,
                    SUM(ei.baik) AS baik,
                    SUM(ei.rusak_ringan) AS rusak_ringan,
                    SUM(ei.rusak_berat) AS rusak_berat
                FROM equipment_inventories ei
                WHERE ei.owner_type = 'App\\Models\\SatkerMabes' AND ei.owner_id IN ({satker_ids_str})
                GROUP BY ei.equipment_id, ei.owner_id
            ) AS inv ON e.id = inv.equipment_id AND rs.id = inv.owner_id
            WHERE e.deleted_at IS NULL
            ORDER BY et.id, e."order", rs.id;
        """
        df_inventory = pd.read_sql(inventory_query, engine)

        wb = Workbook()
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])
        
        for sheet_satker in all_related_satkers:
            sheet_id = sheet_satker['id']  # 🔧 FIX: Gunakan ID
            sheet_name = sheet_satker['name']
            
            # 🔧 FIX: Filter berdasarkan satker_id, bukan satker_name
            df_sheet_data = df_inventory[df_inventory['satker_id'] == sheet_id].copy()
            
            # Buat sheet baru tanpa syarat
            ws = wb.create_sheet(sanitize_name(sheet_name)[:31])
            
            header = ["No.", "Jenis Materil", "Baik", "Rusak Ringan", "Rusak Berat", "Jumlah"]
            ws.append(header)
            
            current_row = 2
            if not df_sheet_data.empty:
                for penggolongan, group_df in df_sheet_data.groupby("penggolongan", sort=False):
                    ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=len(header))
                    ws.cell(row=current_row, column=2, value=penggolongan).font = Font(bold=True)
                    current_row += 1
                    
                    for jenis_no, (_, row) in enumerate(group_df.iterrows(), start=1):
                        baik = int(row["baik"])
                        rr = int(row["rusak_ringan"])
                        rb = int(row["rusak_berat"])
                        jumlah = baik + rr + rb
                        
                        row_data = [
                            jenis_no, 
                            row["jenis_materiil"], 
                            zero_to_empty(baik), 
                            zero_to_empty(rr), 
                            zero_to_empty(rb), 
                            zero_to_empty(jumlah)
                        ]
                        ws.append(row_data)
                        ws.cell(row=current_row, column=1).alignment = Alignment(horizontal="center", vertical="center")
                        current_row += 1
            
            style_header_simple(ws)
            auto_resize_columns(ws)

        # Simpan file tanpa syarat
        print(file_display_name)
        filename = os.path.join(satker_output_dir, f"{sanitize_name(file_display_name)}.xlsx")
        wb.save(filename)
        print(f"    ✅ Saved: {filename}")
            
    print("✅ Satker Mabes export selesai!\n")


# =========================================================
# 3️⃣ EXPORT POLDA, POLRES, POLSEK (Kode Original, tidak diubah)
# =========================================================
if export_polda or export_polres or export_polsek:
    poldas = pd.read_sql("SELECT id, name FROM polda ORDER BY id", engine)
    
    for _, polda in poldas.iterrows():
        polda_id = polda["id"]
        polda_name = polda["name"]
        
        print(f"🚀 Processing POLDA: {polda_name}")
        
        polda_output_dir = os.path.join(output_dir, 'POLDA ' + polda_name)
        polsek_output_dir = os.path.join(polda_output_dir, f"Jajaran Polsek POLDA {polda_name}")
        
        if export_polda:
            os.makedirs(polda_output_dir, exist_ok=True)
        if export_polsek:
            os.makedirs(polsek_output_dir, exist_ok=True)
        
        subsatkers_list_query = f"SELECT name FROM subsatker_poldas WHERE polda_id = {polda_id} ORDER BY name;"
        df_subsatkers_list = pd.read_sql(subsatkers_list_query, engine)
        subsatkers = df_subsatkers_list["name"].tolist()
        
        polres_list_query = f"SELECT id AS polres_id, name AS polres_name FROM polres WHERE polda_id = {polda_id} ORDER BY name;"
        df_polres_list = pd.read_sql(polres_list_query, engine)
        
        if export_polda:
            def style_header_polda(ws):
                for row_num in [1, 2]:
                    for cell in ws[row_num]:
                        cell.font = Font(bold=True)
                        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                ws.freeze_panes = "C3"

            subsatker_query = f"""
                SELECT
                    et.id AS penggolongan_id, et.name AS penggolongan,
                    e.name AS jenis_materiil, e."order",
                    inv.subsatker_name,
                    COALESCE(inv.baik, 0) AS baik,
                    COALESCE(inv.rusak_ringan, 0) AS rusak_ringan,
                    COALESCE(inv.rusak_berat, 0) AS rusak_berat
                FROM equipments e
                JOIN equipment_types et ON et.id = e.id_equipment_type
                LEFT JOIN (
                    SELECT
                        ei.equipment_id, sp.name AS subsatker_name,
                        SUM(ei.baik) AS baik, SUM(ei.rusak_ringan) AS rusak_ringan, SUM(ei.rusak_berat) AS rusak_berat
                    FROM equipment_inventories ei
                    JOIN subsatker_poldas sp ON sp.id = ei.owner_id
                    WHERE ei.owner_type = 'App\\Models\\SubsatkerPolda' AND sp.polda_id = {polda_id}
                    GROUP BY ei.equipment_id, sp.name
                ) AS inv ON e.id = inv.equipment_id
                WHERE e.deleted_at is null
                ORDER BY et.id, e."order";
            """
            df_subsatker = pd.read_sql(subsatker_query, engine)
            
            wb_polda = Workbook()
            ws_polda = wb_polda.active
            ws_polda.title = sanitize_name('POLDA ' + polda_name)
            
            if not df_subsatker.empty:
                header1 = ["No.", "Jenis Materil"]
                for s in subsatkers:
                    header1 += [s, "", "", ""]
                ws_polda.append(header1)
                
                for i, _ in enumerate(subsatkers):
                    start_col = 3 + (i * 4)
                    ws_polda.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=start_col + 3)
                
                header2 = ["", ""]
                header2 += ["Baik", "Rusak Ringan", "Rusak Berat", "Jumlah"] * len(subsatkers)
                ws_polda.append(header2)
                
                current_row = 3
                for penggolongan, group_df in df_subsatker.groupby("penggolongan", sort=False):
                    ws_polda.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=len(header1))
                    ws_polda.cell(row=current_row, column=2, value=penggolongan).font = Font(bold=True)
                    current_row += 1
                    
                    for jenis_no, (jenis, jenis_df) in enumerate(group_df.groupby("jenis_materiil", sort=False), start=1):
                        row_data = [jenis_no, jenis]
                        for s in subsatkers:
                            row = jenis_df[jenis_df["subsatker_name"] == s]
                            baik = int(row["baik"].iloc[0]) if not row.empty else 0
                            rr = int(row["rusak_ringan"].iloc[0]) if not row.empty else 0
                            rb = int(row["rusak_berat"].iloc[0]) if not row.empty else 0
                            
                            jumlah_subsatker = baik + rr + rb
                            row_data += [zero_to_empty(baik), zero_to_empty(rr), zero_to_empty(rb), zero_to_empty(jumlah_subsatker)]
                        
                        ws_polda.append(row_data)
                        ws_polda.cell(row=current_row, column=1).alignment = Alignment(horizontal="center", vertical="center")
                        current_row += 1
                
                style_header_polda(ws_polda)
                auto_resize_columns(ws_polda)
            
            if export_polres:
                for _, polres_row in df_polres_list.iterrows():
                    polres_id = polres_row["polres_id"]
                    polres_name = polres_row["polres_name"]
                    
                    polres_query = f"""
                        SELECT et.name AS penggolongan, e.name AS jenis_materiil,
                               COALESCE(inv.baik, 0) AS baik, COALESCE(inv.rusak_ringan, 0) AS rusak_ringan,
                               COALESCE(inv.rusak_berat, 0) AS rusak_berat
                        FROM equipments e
                        JOIN equipment_types et ON et.id = e.id_equipment_type
                        LEFT JOIN (
                            SELECT ei.equipment_id, SUM(ei.baik) AS baik, SUM(ei.rusak_ringan) AS rusak_ringan, SUM(ei.rusak_berat) AS rusak_berat
                            FROM equipment_inventories ei
                            WHERE ei.owner_type = 'App\\Models\\Polres' AND ei.owner_id = {polres_id}
                            GROUP BY ei.equipment_id
                        ) AS inv ON e.id = inv.equipment_id
                        WHERE e.deleted_at is null
                        ORDER BY et.id, e."order";
                    """
                    df_polres = pd.read_sql(polres_query, engine)
                    
                    if df_polres.empty: continue
                    
                    ws_polres = wb_polda.create_sheet(sanitize_name(polres_name))
                    header_polres = ["No.", "Jenis Materil", "Baik", "Rusak Ringan", "Rusak Berat", "Jumlah"]
                    ws_polres.append(header_polres)
                    
                    current_row = 2
                    for penggolongan, group_df in df_polres.groupby("penggolongan", sort=False):
                        ws_polres.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=len(header_polres))
                        ws_polres.cell(row=current_row, column=2, value=penggolongan).font = Font(bold=True)
                        current_row += 1
                        
                        for jenis_no, (_, row) in enumerate(group_df.iterrows(), start=1):
                            baik = int(row["baik"])
                            rr = int(row["rusak_ringan"])
                            rb = int(row["rusak_berat"])
                            jumlah = baik + rr + rb
                            row_data = [jenis_no, row["jenis_materiil"], zero_to_empty(baik), zero_to_empty(rr), zero_to_empty(rb), zero_to_empty(jumlah)]
                            ws_polres.append(row_data)
                            ws_polres.cell(row=current_row, column=1).alignment = Alignment(horizontal="center", vertical="center")
                            current_row += 1
                    
                    style_header_simple(ws_polres)
                    auto_resize_columns(ws_polres)
            
            polda_filename = os.path.join(polda_output_dir, f"Inventaris_POLDA_{polda_name}.xlsx")
            wb_polda.save(polda_filename)
            print(f"✅ Saved {polda_filename}")
        
        if export_polsek:
            for _, polres_row in df_polres_list.iterrows():
                polres_id = polres_row["polres_id"]
                polres_name = polres_row["polres_name"]
                
                polsek_list_query = f"SELECT id, name FROM polsek WHERE polres_id = {polres_id} ORDER BY name;"
                df_polsek_list = pd.read_sql(polsek_list_query, engine)
                
                if df_polsek_list.empty:
                    continue
                
                wb_polsek = Workbook()
                if "Sheet" in wb_polsek.sheetnames:
                    wb_polsek.remove(wb_polsek["Sheet"])
                
                print(f"  -> Processing Jajaran Polsek untuk POLRES: {polres_name}")
                
                for _, polsek_row in df_polsek_list.iterrows():
                    polsek_id = polsek_row["id"]
                    polsek_name = polsek_row["name"]
                    
                    polsek_query = f"""
                        SELECT et.name AS penggolongan, e.name AS jenis_materiil,
                               COALESCE(inv.baik, 0) AS baik, COALESCE(inv.rusak_ringan, 0) AS rusak_ringan,
                               COALESCE(inv.rusak_berat, 0) AS rusak_berat
                        FROM equipments e
                        JOIN equipment_types et ON et.id = e.id_equipment_type
                        LEFT JOIN (
                            SELECT ei.equipment_id, SUM(ei.baik) AS baik, SUM(ei.rusak_ringan) AS rusak_ringan, SUM(ei.rusak_berat) AS rusak_berat
                            FROM equipment_inventories ei
                            WHERE ei.owner_type = 'App\\Models\\Polsek' AND ei.owner_id = {polsek_id}
                            GROUP BY ei.equipment_id
                        ) AS inv ON e.id = inv.equipment_id
                        WHERE e.deleted_at is null
                        ORDER BY et.id, e."order";
                    """
                    df_polsek = pd.read_sql(polsek_query, engine)
                    
                    if sum(df_polsek['baik']) + sum(df_polsek['rusak_ringan']) + sum(df_polsek['rusak_berat']) == 0:
                        continue
                    
                    ws_polsek = wb_polsek.create_sheet(sanitize_name(polsek_name))
                    header_polsek = ["No.", "Jenis Materil", "Baik", "Rusak Ringan", "Rusak Berat", "Jumlah"]
                    ws_polsek.append(header_polsek)
                    
                    current_row = 2
                    for penggolongan, group_df in df_polsek.groupby("penggolongan", sort=False):
                        ws_polsek.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=len(header_polsek))
                        ws_polsek.cell(row=current_row, column=2, value=penggolongan).font = Font(bold=True)
                        current_row += 1
                        
                        for jenis_no, (_, row) in enumerate(group_df.iterrows(), start=1):
                            baik = int(row["baik"])
                            rr = int(row["rusak_ringan"])
                            rb = int(row["rusak_berat"])
                            jumlah = baik + rr + rb
                            row_data = [jenis_no, row["jenis_materiil"], zero_to_empty(baik), zero_to_empty(rr), zero_to_empty(rb), zero_to_empty(jumlah)]
                            ws_polsek.append(row_data)
                            ws_polsek.cell(row=current_row, column=1).alignment = Alignment(horizontal="center", vertical="center")
                            current_row += 1
                    
                    style_header_simple(ws_polsek)
                    auto_resize_columns(ws_polsek)
                
                if len(wb_polsek.sheetnames) > 0:
                    polsek_filename = os.path.join(polsek_output_dir, f"Inventaris_Polsek_{polres_name}.xlsx")
                    wb_polsek.save(polsek_filename)
                    print(f"  ✅ Saved Jajaran Polsek: {polsek_filename}")

# =========================================================
# 4️⃣ EXPORT SATKER MABES
# =========================================================
if export_satker_mabes:
    export_satker_mabes()

print("\n🎉 Semua file selesai dibuat di folder 'exports'!")