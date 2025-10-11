import os
import pandas as pd
import psycopg2
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from sqlalchemy import create_engine

# =========================================================
# 1️⃣ Load konfigurasi dari .env
# =========================================================
load_dotenv()

DB_HOST = os.getenv("DB_HOST")
DB_PORT = os.getenv("DB_PORT")
DB_NAME = os.getenv("DB_DATABASE")
DB_USER = os.getenv("DB_USERNAME")
DB_PASS = os.getenv("DB_PASSWORD")

engine = create_engine(f"postgresql://{DB_USER}:{DB_PASS}@{DB_HOST}:{DB_PORT}/{DB_NAME}")

# =========================================================
# 2️⃣ Ambil daftar semua POLDA
# =========================================================
poldas = pd.read_sql("SELECT id, name FROM polda ORDER BY id", engine)

# Direktori output utama
output_dir = "exports"
os.makedirs(output_dir, exist_ok=True)

# =========================================================
# 3️⃣ Fungsi bantu buat format header merge cell
# =========================================================
def style_header(ws):
    # Styling untuk baris header pertama dan kedua
    for row_num in [1, 2]:
        for cell in ws[row_num]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = "C3" # Freeze di bawah header

# =========================================================
# 4️⃣ Fungsi untuk auto-resize kolom
# =========================================================
def auto_resize_columns(ws):
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        
        for cell in column:
            try:
                if cell.value:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
            except:
                pass
        
        adjusted_width = min(max_length + 2, 50)  # Max 50 karakter
        ws.column_dimensions[column_letter].width = adjusted_width

# =========================================================
# 5️⃣ Fungsi bantu untuk mengganti 0 menjadi string kosong
# =========================================================
def zero_to_empty(value):
    """Mengubah nilai 0 menjadi string kosong."""
    return "" if value == 0 else value
    
# =========================================================
# 6️⃣ Fungsi bantu untuk membersihkan nama sheet/file
# =========================================================
def sanitize_name(name):
    """Membersihkan nama untuk digunakan sebagai nama file atau sheet."""
    return name[:31].replace('/', '-').replace('\\', '-').replace('*', '').replace('?', '').replace(':', '').replace('[', '').replace(']', '')

# =========================================================
# 7️⃣ Loop tiap POLDA
# =========================================================
for _, polda in poldas.iterrows():
    polda_id = polda["id"]
    polda_name = polda["name"]

    print(f"🚀 Processing POLDA: {polda_name}")

    # === PERUBAHAN STRUKTUR FOLDER ===
    polda_output_dir = os.path.join(output_dir, 'POLDA ' + polda_name)
    polsek_output_dir = os.path.join(polda_output_dir, f"Jajaran Polsek POLDA {polda_name}")
    os.makedirs(polda_output_dir, exist_ok=True)
    os.makedirs(polsek_output_dir, exist_ok=True)
    
    # === Ambil daftar Subsatker untuk POLDA ini (untuk header kolom) ===
    subsatkers_list_query = f"SELECT name FROM subsatker_poldas WHERE polda_id = {polda_id} ORDER BY name;"
    df_subsatkers_list = pd.read_sql(subsatkers_list_query, engine)
    subsatkers = df_subsatkers_list["name"].tolist()

    # === Query untuk Sheet 1: Subsatker POLDA ===
    subsatker_query = f"""
        SELECT
            et.id AS penggolongan_id, et.name AS penggolongan,
            e.name AS jenis_materiil, e."order",
            inv.subsatker_name,
            COALESCE(inv.baik, 0) AS baik,
            COALESCE(inv.rusak_ringan, 0) AS rusak_ringan,
            COALESCE(inv.rusak_berat, 0) AS rusak_berat
        FROM equipments e
        JOIN equipment_types et ON et.id = e.id_equipment_type
        LEFT JOIN (
            SELECT
                ei.equipment_id, sp.name AS subsatker_name,
                SUM(ei.baik) AS baik, SUM(ei.rusak_ringan) AS rusak_ringan, SUM(ei.rusak_berat) AS rusak_berat
            FROM equipment_inventories ei
            JOIN subsatker_poldas sp ON sp.id = ei.owner_id
            WHERE ei.owner_type = 'App\\Models\\SubsatkerPolda' AND sp.polda_id = {polda_id}
            GROUP BY ei.equipment_id, sp.name
        ) AS inv ON e.id = inv.equipment_id
        WHERE e.deleted_at is null
        ORDER BY et.id, e."order";
    """
    df_subsatker = pd.read_sql(subsatker_query, engine)

    # === Ambil semua POLRES untuk POLDA ini ===
    polres_list_query = f"SELECT id AS polres_id, name AS polres_name FROM polres WHERE polda_id = {polda_id} ORDER BY name;"
    df_polres_list = pd.read_sql(polres_list_query, engine)

    # =========================================================
    # 8️⃣ Buat file Excel utama untuk POLDA
    # =========================================================
    wb_polda = Workbook()
    ws_polda = wb_polda.active
    ws_polda.title = sanitize_name('POLDA ' + polda_name)

    # ===================== SHEET POLDA =====================
    if not df_subsatker.empty:
        header1 = ["No.", "Jenis Materil"]
        for s in subsatkers:
            header1 += [s, "", "", ""] 
        ws_polda.append(header1)
        
        # Merge cells untuk header Subsatker
        for i, _ in enumerate(subsatkers):
            start_col = 3 + (i * 4) 
            ws_polda.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=start_col + 3)

        header2 = ["", ""]
        header2 += ["Baik", "Rusak Ringan", "Rusak Berat", "Jumlah"] * len(subsatkers)
        ws_polda.append(header2)
        
        current_row = 3
        for penggolongan, group_df in df_subsatker.groupby("penggolongan", sort=False):
            ws_polda.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=len(header1))
            ws_polda.cell(row=current_row, column=2, value=penggolongan).font = Font(bold=True)
            current_row += 1

            for jenis_no, (jenis, jenis_df) in enumerate(group_df.groupby("jenis_materiil", sort=False), start=1):
                row_data = [jenis_no, jenis]
                for s in subsatkers:
                    row = jenis_df[jenis_df["subsatker_name"] == s]
                    baik = int(row["baik"].iloc[0]) if not row.empty else 0
                    rr = int(row["rusak_ringan"].iloc[0]) if not row.empty else 0
                    rb = int(row["rusak_berat"].iloc[0]) if not row.empty else 0
                    
                    jumlah_subsatker = baik + rr + rb
                    row_data += [zero_to_empty(baik), zero_to_empty(rr), zero_to_empty(rb), zero_to_empty(jumlah_subsatker)]

                ws_polda.append(row_data)
                ws_polda.cell(row=current_row, column=1).alignment = Alignment(horizontal="center", vertical="center")
                current_row += 1
        
        style_header(ws_polda)
        auto_resize_columns(ws_polda)

    # ===================== SHEET PER POLRES (di file utama Polda) =====================
    for _, polres_row in df_polres_list.iterrows():
        polres_id = polres_row["polres_id"]
        polres_name = polres_row["polres_name"]
        
        polres_query = f"""
            SELECT et.name AS penggolongan, e.name AS jenis_materiil,
                   COALESCE(inv.baik, 0) AS baik, COALESCE(inv.rusak_ringan, 0) AS rusak_ringan,
                   COALESCE(inv.rusak_berat, 0) AS rusak_berat
            FROM equipments e
            JOIN equipment_types et ON et.id = e.id_equipment_type
            LEFT JOIN (
                SELECT ei.equipment_id, SUM(ei.baik) AS baik, SUM(ei.rusak_ringan) AS rusak_ringan, SUM(ei.rusak_berat) AS rusak_berat
                FROM equipment_inventories ei
                WHERE ei.owner_type = 'App\\Models\\Polres' AND ei.owner_id = {polres_id}
                GROUP BY ei.equipment_id
            ) AS inv ON e.id = inv.equipment_id
            WHERE e.deleted_at is null
            ORDER BY et.id, e."order";
        """
        df_polres = pd.read_sql(polres_query, engine)
        
        if df_polres.empty: continue
            
        ws_polres = wb_polda.create_sheet(sanitize_name(polres_name))
        header_polres = ["No.", "Jenis Materil", "Baik", "Rusak Ringan", "Rusak Berat", "Jumlah"]
        ws_polres.append(header_polres)
        
        current_row = 2
        for penggolongan, group_df in df_polres.groupby("penggolongan", sort=False):
            ws_polres.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=len(header_polres))
            ws_polres.cell(row=current_row, column=2, value=penggolongan).font = Font(bold=True)
            current_row += 1

            for jenis_no, (_, row) in enumerate(group_df.iterrows(), start=1):
                baik = int(row["baik"])
                rr = int(row["rusak_ringan"])
                rb = int(row["rusak_berat"])
                jumlah = baik + rr + rb
                row_data = [jenis_no, row["jenis_materiil"], zero_to_empty(baik), zero_to_empty(rr), zero_to_empty(rb), zero_to_empty(jumlah)]
                ws_polres.append(row_data)
                ws_polres.cell(row=current_row, column=1).alignment = Alignment(horizontal="center", vertical="center")
                current_row += 1
        
        for cell in ws_polres[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
        ws_polres.freeze_panes = "A2"
        auto_resize_columns(ws_polres)

    # Simpan file utama POLDA
    polda_filename = os.path.join(polda_output_dir, f"Inventaris_POLDA_{polda_name}.xlsx")
    wb_polda.save(polda_filename)
    print(f"✅ Saved {polda_filename}")

    # =========================================================
    # 9️⃣ BAGIAN BARU: Buat file terpisah untuk jajaran POLSEK
    # =========================================================
    for _, polres_row in df_polres_list.iterrows():
        polres_id = polres_row["polres_id"]
        polres_name = polres_row["polres_name"]
        
        # Ambil daftar polsek di bawah polres ini
        polsek_list_query = f"SELECT id, name FROM polsek WHERE polres_id = {polres_id} ORDER BY name;"
        df_polsek_list = pd.read_sql(polsek_list_query, engine)

        if df_polsek_list.empty:
            continue

        # Buat workbook baru untuk setiap Polres yang berisi Polsek
        wb_polsek = Workbook()
        # Hapus sheet default yang dibuat otomatis
        if "Sheet" in wb_polsek.sheetnames:
            wb_polsek.remove(wb_polsek["Sheet"])

        print(f"  -> Processing Jajaran Polsek untuk POLRES: {polres_name}")

        for _, polsek_row in df_polsek_list.iterrows():
            polsek_id = polsek_row["id"]
            polsek_name = polsek_row["name"]

            # Query data inventaris untuk polsek ini
            polsek_query = f"""
                SELECT et.name AS penggolongan, e.name AS jenis_materiil,
                       COALESCE(inv.baik, 0) AS baik, COALESCE(inv.rusak_ringan, 0) AS rusak_ringan,
                       COALESCE(inv.rusak_berat, 0) AS rusak_berat
                FROM equipments e
                JOIN equipment_types et ON et.id = e.id_equipment_type
                LEFT JOIN (
                    SELECT ei.equipment_id, SUM(ei.baik) AS baik, SUM(ei.rusak_ringan) AS rusak_ringan, SUM(ei.rusak_berat) AS rusak_berat
                    FROM equipment_inventories ei
                    WHERE ei.owner_type = 'App\\Models\\Polsek' AND ei.owner_id = {polsek_id}
                    GROUP BY ei.equipment_id
                ) AS inv ON e.id = inv.equipment_id
                WHERE e.deleted_at is null
                ORDER BY et.id, e."order";
            """
            df_polsek = pd.read_sql(polsek_query, engine)
            
            # Buat sheet hanya jika ada data inventaris
            if sum(df_polsek['baik']) + sum(df_polsek['rusak_ringan']) + sum(df_polsek['rusak_berat']) == 0:
                continue

            ws_polsek = wb_polsek.create_sheet(sanitize_name(polsek_name))
            header_polsek = ["No.", "Jenis Materil", "Baik", "Rusak Ringan", "Rusak Berat", "Jumlah"]
            ws_polsek.append(header_polsek)

            current_row = 2
            for penggolongan, group_df in df_polsek.groupby("penggolongan", sort=False):
                ws_polsek.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=len(header_polsek))
                ws_polsek.cell(row=current_row, column=2, value=penggolongan).font = Font(bold=True)
                current_row += 1

                for jenis_no, (_, row) in enumerate(group_df.iterrows(), start=1):
                    baik = int(row["baik"])
                    rr = int(row["rusak_ringan"])
                    rb = int(row["rusak_berat"])
                    jumlah = baik + rr + rb
                    row_data = [jenis_no, row["jenis_materiil"], zero_to_empty(baik), zero_to_empty(rr), zero_to_empty(rb), zero_to_empty(jumlah)]
                    ws_polsek.append(row_data)
                    ws_polsek.cell(row=current_row, column=1).alignment = Alignment(horizontal="center", vertical="center")
                    current_row += 1

            for cell in ws_polsek[1]:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center")
            ws_polsek.freeze_panes = "A2"
            auto_resize_columns(ws_polsek)
        
        # Simpan file polsek jika ada sheet yang dibuat
        if len(wb_polsek.sheetnames) > 0:
            polsek_filename = os.path.join(polsek_output_dir, f"Inventaris_Polsek_{polres_name}.xlsx")
            wb_polsek.save(polsek_filename)
            print(f"  ✅ Saved Jajaran Polsek: {polsek_filename}")

print("\n🎉 Semua file selesai dibuat di folder 'exports' dengan struktur baru!")